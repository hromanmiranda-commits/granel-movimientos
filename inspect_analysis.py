import openpyxl

wb = openpyxl.load_workbook("Ventas Granel.xlsx", data_only=False)

if "Análisis de Negocio" in wb.sheetnames:
    ws = wb["Análisis de Negocio"]
    print("=== Análisis de Negocio ===")
    for row in range(1, 20):
        print(f"{ws.cell(row, 1).value} : {ws.cell(row, 2).value}")

if "Ignacio" in wb.sheetnames:
    ws = wb["Ignacio"]
    print("\n=== Ignacio ===")
    for row in range(4, 10):
        for col in range(1, 8):
            val = ws.cell(row, col).value
            if val is not None:
                print(f"R{row}C{col}: {val}")

if "ENAP carga" in wb.sheetnames:
    ws = wb["ENAP carga"]
    print("\n=== ENAP carga ===")
    for row in range(1, 10):
        row_vals = []
        for col in range(1, 10):
             row_vals.append(str(ws.cell(row, col).value))
        print(" | ".join(row_vals))

