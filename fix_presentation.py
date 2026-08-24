import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

file_path = "Ventas Granel.xlsx"
wb = openpyxl.load_workbook(file_path)

bold_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")

def format_header(cell):
    cell.font = bold_font
    cell.fill = header_fill
    cell.alignment = align_center

def format_value(cell):
    cell.alignment = align_center

def clear_rows(ws, row_start, row_end, col_start, col_end):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)
            ws.cell(row=r, column=c).font = Font()

# 1. Update 'Ignacio' tab
if "Ignacio" in wb.sheetnames:
    ws_ignacio = wb["Ignacio"]
    clear_rows(ws_ignacio, 5, 6, 2, 8)
    
    headers = ["LITROS VENDIDOS", "EXTRACCIÓN (L)", "TOTAL VENTAS ($)", "N° OPERACIONES"]
    formulas = [
        "=SUMIFS('Detalles movimientos'!I$5:I$1000, 'Detalles movimientos'!F$5:F$1000, \"Ignacio\")",
        "=SUMIFS('Detalles movimientos'!L$5:L$1000, 'Detalles movimientos'!F$5:F$1000, \"Ignacio\")",
        "=SUMIFS('Detalles movimientos'!M$5:M$1000, 'Detalles movimientos'!F$5:F$1000, \"Ignacio\")",
        "=COUNTIFS('Detalles movimientos'!F$5:F$1000, \"Ignacio\")"
    ]
    
    for i, (head, form) in enumerate(zip(headers, formulas)):
        col = i + 2 # Start at column B (2)
        cell_h = ws_ignacio.cell(row=5, column=col)
        cell_h.value = head
        format_header(cell_h)
        ws_ignacio.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        cell_v = ws_ignacio.cell(row=6, column=col)
        cell_v.value = form
        format_value(cell_v)

# 2. Update 'ENAP' tab
if "ENAP" in wb.sheetnames:
    ws_enap = wb["ENAP"]
    clear_rows(ws_enap, 5, 6, 2, 8)
    
    headers = ["LITROS VENDIDOS", "LITROS COMPRADOS", "STOCK ACTUAL", "TOTAL VENTAS ($)", "N° OPERACIONES"]
    formulas = [
        "=SUMIFS('Detalles movimientos'!I$5:I$1000, 'Detalles movimientos'!F$5:F$1000, \"ENAP carga\")",
        "=SUMIFS('Detalles movimientos'!R$5:R$1000, 'Detalles movimientos'!F$5:F$1000, \"ENAP carga\")",
        "=C6-B6",
        "=SUMIFS('Detalles movimientos'!M$5:M$1000, 'Detalles movimientos'!F$5:F$1000, \"ENAP carga\")",
        "=COUNTIFS('Detalles movimientos'!F$5:F$1000, \"ENAP carga\")"
    ]
    
    for i, (head, form) in enumerate(zip(headers, formulas)):
        col = i + 2 # Start at column B (2)
        cell_h = ws_enap.cell(row=5, column=col)
        cell_h.value = head
        format_header(cell_h)
        ws_enap.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        cell_v = ws_enap.cell(row=6, column=col)
        cell_v.value = form
        format_value(cell_v)

wb.save(file_path)
print("Fixes applied successfully.")
