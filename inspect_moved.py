import pandas as pd
import openpyxl

file_path = "Ventas Granel.xlsx"
xls = pd.ExcelFile(file_path)

print("=== Detalles movimientos ===")
df = pd.read_excel(xls, sheet_name='Detalles movimientos', header=None, nrows=15)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
print(df)

wb = openpyxl.load_workbook(file_path, data_only=False)
for sheet_name in ["Análisis de Negocio", "Ignacio", "ENAP"]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===")
        for r in range(1, 15):
            row_vals = []
            for c in range(1, 6):
                val = ws.cell(row=r, column=c).value
                row_vals.append(str(val) if val is not None else "")
            if any(row_vals):
                print(f"Row {r}: " + " | ".join(row_vals))
