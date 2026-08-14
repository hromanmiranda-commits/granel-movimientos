import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

file_path = "Ventas Granel.xlsx"
wb = openpyxl.load_workbook(file_path)

bold_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
align_center = Alignment(horizontal="center")

# 1. Update 'Ignacio' tab
if "Ignacio" in wb.sheetnames:
    ws_ignacio = wb["Ignacio"]
    
    # Add TOTAL EXTRACCIÓN in C5 and C6
    ws_ignacio['C5'] = "TOTAL EXTRACCIÓN (L)"
    ws_ignacio['C5'].font = bold_font
    ws_ignacio['C5'].fill = header_fill
    ws_ignacio['C5'].alignment = align_center
    
    # Formula for sum of extraction (Column J in 'Detalles movimientos')
    ws_ignacio['C6'] = "=SUMIFS('Detalles movimientos'!J$5:J$1000, 'Detalles movimientos'!F$5:F$1000, \"Ignacio\")"
    
# 2. Update 'ENAP' tab
if "ENAP" in wb.sheetnames:
    ws_enap = wb["ENAP"]
    
    # Rename B5 to LITROS VENDIDOS
    ws_enap['B5'] = "LITROS VENDIDOS"
    ws_enap['B5'].font = bold_font
    ws_enap['B5'].fill = header_fill
    ws_enap['B5'].alignment = align_center
    
    # B6 is already =SUMIFS(... I$5:I$1000 ...), which is correct for Litros Vendidos
    
    # Add LITROS COMPRADOS in C5
    ws_enap['C5'] = "LITROS COMPRADOS"
    ws_enap['C5'].font = bold_font
    ws_enap['C5'].fill = header_fill
    ws_enap['C5'].alignment = align_center
    
    # Formula for sum of bought (Column P in 'Detalles movimientos')
    ws_enap['C6'] = "=SUMIFS('Detalles movimientos'!P$5:P$1000, 'Detalles movimientos'!F$5:F$1000, \"ENAP carga\")"
    
    # Add STOCK ACTUAL in D5
    ws_enap['D5'] = "STOCK ACTUAL"
    ws_enap['D5'].font = bold_font
    ws_enap['D5'].fill = header_fill
    ws_enap['D5'].alignment = align_center
    
    # Formula for stock (Comprados - Vendidos)
    ws_enap['D6'] = "=C6-B6"

wb.save(file_path)
print("Excel file updated successfully.")
