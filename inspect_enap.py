import openpyxl

wb = openpyxl.load_workbook("Ventas Granel.xlsx", data_only=False)

if "ENAP" in wb.sheetnames:
    ws = wb["ENAP"]
    print("\n=== ENAP ===")
    for row in range(1, 15):
        row_vals = []
        for col in range(1, 10):
             val = ws.cell(row, col).value
             row_vals.append(str(val) if val is not None else "None")
        print(" | ".join(row_vals))

