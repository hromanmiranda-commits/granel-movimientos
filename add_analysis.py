import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

file_path = "Ventas Granel.xlsx"
wb = openpyxl.load_workbook(file_path)

# If the analysis sheet already exists, remove it to start fresh
sheet_name = "Análisis de Negocio"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

# Create the new sheet as the second sheet (index 1)
ws = wb.create_sheet(title=sheet_name, index=1)

# Define styles
header_font = Font(bold=True, size=14, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
bold_font = Font(bold=True)
fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Set title
ws['A1'] = "Dashboard de Análisis de Negocio"
ws.merge_cells('A1:B1')
ws['A1'].font = header_font
ws['A1'].fill = header_fill
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Define labels and formulas
metrics = [
    ("Métricas de Inventario (Litros)", ""),
    ("Total Litros Comprados (ENAP)", "=SUM('Detalles movimientos'!O:O)"),
    ("Total Litros Vendidos", "=SUM('Detalles movimientos'!I:I)"),
    ("Saldo Actual (Stock)", "=B3-B4"),
    ("", ""),
    ("Métricas Financieras ($)", ""),
    ("Monto Total de Compras (ENAP)", "=SUM('Detalles movimientos'!N:N)"),
    ("Total Recaudado (Ventas)", "=SUM('Detalles movimientos'!J:J)"),
    ("Total Comisiones Pagadas", "=SUM('Detalles movimientos'!L:L)"),
    ("Beneficio Bruto", "=B9-B8"),
    ("Beneficio Neto", "=B11-B10"),
    ("", ""),
    ("Indicadores Clave de Rendimiento (KPIs)", ""),
    ("Costo Promedio por Litro", "=IF(B3>0, B8/B3, 0)"),
    ("Precio Promedio de Venta por Litro", "=IF(B4>0, B9/B4, 0)"),
    ("Margen de Ganancia Neto", "=IF(B9>0, B12/B9, 0)")
]

row = 3
for label, formula in metrics:
    ws.cell(row=row, column=1, value=label)
    if formula:
        ws.cell(row=row, column=2, value=formula)
    row += 1

# Apply formatting
for r in range(3, 19):
    # If section header (no formula)
    if not ws.cell(row=r, column=2).value:
        ws.cell(row=r, column=1).font = bold_font
        ws.cell(row=r, column=1).fill = fill_gray
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    else:
        # Format numbers
        cell = ws.cell(row=r, column=2)
        if "Litros" in ws.cell(row=r, column=1).value or "Saldo" in ws.cell(row=r, column=1).value:
            cell.number_format = '#,##0.0'
        elif "Margen" in ws.cell(row=r, column=1).value:
            cell.number_format = '0.00%'
        else:
            cell.number_format = '"$"#,##0'

# Adjust column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 25

wb.save(file_path)
print("Analysis sheet added successfully.")
