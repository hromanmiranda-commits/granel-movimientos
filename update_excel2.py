import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

file_path = "Ventas Granel.xlsx"
wb = openpyxl.load_workbook(file_path)

bold_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")

def format_header(cell):
    cell.font = bold_font
    cell.fill = header_fill
    cell.alignment = align_center

def format_value(cell):
    cell.alignment = align_center

# 1. Update 'Ignacio' tab
if "Ignacio" in wb.sheetnames:
    ws_ignacio = wb["Ignacio"]
    
    if "B5:C5" in [str(r) for r in ws_ignacio.merged_cells.ranges]:
        ws_ignacio.unmerge_cells("B5:C5")
    if "B6:C7" in [str(r) for r in ws_ignacio.merged_cells.ranges]:
        ws_ignacio.unmerge_cells("B6:C7")
        
    ws_ignacio['B5'] = "TOTAL LITROS"
    format_header(ws_ignacio['B5'])
    ws_ignacio['B6'] = "=SUMIFS('Detalles movimientos'!I$5:I$1000, 'Detalles movimientos'!F$5:F$1000, \"Ignacio\")"
    format_value(ws_ignacio['B6'])
    
    ws_ignacio['C5'] = "EXTRACCIÓN (L)"
    format_header(ws_ignacio['C5'])
    ws_ignacio['C6'] = "=SUMIFS('Detalles movimientos'!J$5:J$1000, 'Detalles movimientos'!F$5:F$1000, \"Ignacio\")"
    format_value(ws_ignacio['C6'])
    
    ws_ignacio.column_dimensions['B'].width = 15
    ws_ignacio.column_dimensions['C'].width = 15

# 2. Update 'ENAP' tab
if "ENAP" in wb.sheetnames:
    ws_enap = wb["ENAP"]
    
    if "B5:C5" in [str(r) for r in ws_enap.merged_cells.ranges]:
        ws_enap.unmerge_cells("B5:C5")
    if "B6:C7" in [str(r) for r in ws_enap.merged_cells.ranges]:
        ws_enap.unmerge_cells("B6:C7")
        
    ws_enap['B5'] = "LITROS VENDIDOS"
    format_header(ws_enap['B5'])
    ws_enap['B6'] = "=SUMIFS('Detalles movimientos'!I$5:I$1000, 'Detalles movimientos'!F$5:F$1000, \"ENAP carga\")"
    format_value(ws_enap['B6'])
    
    ws_enap['C5'] = "LITROS COMPRADOS"
    format_header(ws_enap['C5'])
    ws_enap['C6'] = "=SUMIFS('Detalles movimientos'!P$5:P$1000, 'Detalles movimientos'!F$5:F$1000, \"ENAP carga\")"
    format_value(ws_enap['C6'])
    
    ws_enap['D5'] = "STOCK ACTUAL"
    format_header(ws_enap['D5'])
    ws_enap['D6'] = "=C6-B6"
    format_value(ws_enap['D6'])

    ws_enap.column_dimensions['B'].width = 18
    ws_enap.column_dimensions['C'].width = 18
    ws_enap.column_dimensions['D'].width = 15

wb.save(file_path)
print("Excel file updated successfully.")
